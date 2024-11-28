from openpyxl import load_workbook, Workbook

# Fungsi untuk membaca data dari file Excel
def read_excel(Data_Servis):
    try:
        wb = load_workbook('data_servis.xlsx')
        if Data_Servis in wb.sheetnames:
            sheet = wb[Data_Servis]
            data = []
            for row in sheet.iter_rows(min_row=2, values_only=True):  # Mulai dari baris ke-2 (tanpa header)
                data.append(row)
            return data
        else:
            print(f"Sheet '{Data_Servis}' tidak ditemukan!")
            return None
    except FileNotFoundError:
        print("File Excel tidak ditemukan! Pastikan file 'data_servis.xlsx' tersedia.")
        return None

# Fungsi untuk menulis data ke file Excel
def write_excel(Data_Servis, data_row):
    try:
        wb = load_workbook('data_servis.xlsx')
    except FileNotFoundError:
        wb = Workbook()

    if Data_Servis not in wb.sheetnames:
        ws = wb.create_sheet(Data_Servis)
    else:
        ws = wb[Data_Servis]

    ws.append(data_row)
    wb.save('data_servis.xlsx')
